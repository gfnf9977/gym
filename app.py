from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import os

app = Flask(__name__)
CORS(app)

# Use raw string or forward slashes for Windows path
EXCEL_FILE = r"C:\Users\38093\gym\public\gymexp.xlsx"
# Alternative way using forward slashes:
# EXCEL_FILE = "C:/Users/38093/gym/public/gymexp.xlsx"

@app.route('/update-weight', methods=['POST'])
def update_weight():
    try:
        # First verify if file exists
        if not os.path.exists(EXCEL_FILE):
            print(f"Excel file not found at: {EXCEL_FILE}")
            return jsonify({"error": f"Excel file not found at {EXCEL_FILE}"}), 404

        data = request.json
        print(f"Received data: {data}")

        new_weight = data.get('weight')
        username = data.get('username')

        if new_weight is None or username is None:
            return jsonify({"error": "Missing weight or username"}), 400

        print(f"Updating weight for {username} to {new_weight}")

        try:
            workbook = openpyxl.load_workbook(EXCEL_FILE)
        except Exception as e:
            print(f"Error loading workbook: {str(e)}")
            return jsonify({"error": f"Could not load Excel file: {str(e)}"}), 500

        sheet = workbook['Лист1']

        # Find column indices
        header_row = sheet[1]
        username_col = None
        weight_col = None

        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'username':
                username_col = idx
            elif cell.value == 'weight':
                weight_col = idx

        if not username_col or not weight_col:
            return jsonify({"error": "Column headers not found"}), 400

        # Find and update the user's weight
        user_found = False
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=username_col).value == username:
                sheet.cell(row=row, column=weight_col).value = float(new_weight)
                user_found = True
                break

        if not user_found:
            return jsonify({"error": "Username not found"}), 404

        # Save the workbook
        try:
            workbook.save(EXCEL_FILE)
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            return jsonify({"error": f"Could not save Excel file: {str(e)}"}), 500

        return jsonify({"message": "Weight updated successfully"}), 200

    except Exception as e:
        print(f"Error in update_weight: {str(e)}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    # Print the full path at startup to verify
    print(f"Looking for Excel file at: {os.path.abspath(EXCEL_FILE)}")
    app.run(debug=True, port=5000)